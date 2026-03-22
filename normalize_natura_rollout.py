import argparse
import csv
import re
import sys
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:
    print("ERROR: Falta dependencia openpyxl. Instala con: pip install openpyxl")
    sys.exit(1)


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_email(value: str) -> str:
    return normalize_spaces(value).lower()


def normalize_phone(value: str) -> str:
    raw = normalize_spaces(value)
    if not raw:
        return ""
    raw = re.sub(r"ext.*$", "", raw, flags=re.IGNORECASE).strip()
    plus = raw.startswith("+")
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return ""
    return f"+{digits}" if plus else digits


def title_case(value: str) -> str:
    text = normalize_spaces(value)
    if not text:
        return ""
    return " ".join(part.capitalize() for part in text.split(" "))


def parse_home_address(address: str) -> tuple[str, str, str, str]:
    lines = [normalize_spaces(line) for line in (address or "").splitlines() if normalize_spaces(line)]
    if not lines:
        return "", "", "", ""

    first = lines[0]
    floor = ""
    apartment = ""

    for line in lines[1:]:
        lower = line.lower()
        if "piso" in lower and not floor:
            floor_match = re.search(r"piso\s*([^\s,]+)", line, flags=re.IGNORECASE)
            floor = normalize_spaces(floor_match.group(1) if floor_match else line)
        if ("departamento" in lower or "depto" in lower or "dpto" in lower) and not apartment:
            apt_match = re.search(
                r"(departamento|depto|dpto)\s*([^\s,]+)",
                line,
                flags=re.IGNORECASE,
            )
            apartment = normalize_spaces(apt_match.group(2) if apt_match else line)

    if re.search(r"\bs/?n\b", first, flags=re.IGNORECASE):
        return first, "S/N", floor, apartment

    match = re.match(r"^(.*?)[,\s]+(\d+[A-Za-z0-9/-]*)$", first)
    if match:
        street = normalize_spaces(match.group(1))
        number = normalize_spaces(match.group(2))
        return street, number, floor, apartment

    return first, "", floor, apartment


def get_value(row: dict[str, Any], *candidates: str) -> str:
    lowered = {str(k).strip().lower(): row[k] for k in row.keys()}
    for candidate in candidates:
        key = candidate.strip().lower()
        if key in lowered:
            value = lowered[key]
            return "" if value is None else str(value)
    return ""


def run(input_path: str, normalized_path: str, rejected_path: str) -> None:
    wb = load_workbook(filename=input_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("El archivo no tiene filas.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]

    normalized: list[dict[str, str]] = []
    rejected: list[dict[str, str]] = []

    def value_by_index(values_tuple: tuple[Any, ...], index: int) -> str:
        if index < len(values_tuple) and values_tuple[index] is not None:
            return str(values_tuple[index])
        return ""

    for idx, values in enumerate(data_rows, start=2):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}

        display_name = title_case(get_value(row, "Nombre legal"))
        email = normalize_email(get_value(row, "Work email adress", "Work email address"))
        cp = normalize_spaces(get_value(row, "Home Address Post Code"))
        locality = title_case(get_value(row, "Ciudad"))
        county_raw = normalize_spaces(get_value(row, "County"))
        province_guess = re.sub(r"^\d+\s*", "", county_raw).strip()
        province = title_case(province_guess if province_guess else locality)
        address = get_value(row, "Home Address")
        phone = normalize_phone(get_value(row, "Phone", "Telefono", "Teléfono", "Mobile"))
        if not phone:
            phone = normalize_phone(value_by_index(values, 9))

        street, number, floor, apartment = parse_home_address(address)

        errors = []
        if not display_name:
            errors.append("Nombre vacio")
        if not email or "@" not in email:
            errors.append("Email invalido")
        if not street:
            errors.append("Calle vacia")
        if not cp:
            errors.append("CP vacio")
        if not locality:
            errors.append("Ciudad vacia")

        notes = ""
        if not number:
            notes = "Numero no detectado automaticamente"

        if errors:
            rejected.append(
                {
                    "row": str(idx),
                    "display_name": display_name,
                    "email": email,
                    "reason": "; ".join(errors),
                }
            )
            continue

        normalized.append(
            {
                "display_name": display_name,
                "email": email,
                "phone": phone,
                "street": street,
                "number": number or "S/N",
                "floor": floor,
                "apartment": apartment,
                "locality": locality,
                "province": province,
                "cp": cp,
                "notes": notes,
                "is_active": "1",
            }
        )

    seen = set()
    deduped: list[dict[str, str]] = []
    for item in normalized:
        key = item["email"]
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)

    with open(normalized_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "display_name",
                "email",
                "phone",
                "street",
                "number",
                "floor",
                "apartment",
                "locality",
                "province",
                "cp",
                "notes",
                "is_active",
            ],
        )
        writer.writeheader()
        writer.writerows(deduped)

    with open(rejected_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["row", "display_name", "email", "reason"])
        writer.writeheader()
        writer.writerows(rejected)

    print(f"OK: normalizados={len(deduped)} rechazados={len(rejected)}")
    print(f"Archivo normalizado: {normalized_path}")
    print(f"Archivo rechazados: {rejected_path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--normalized", default="quick_users_normalizado.csv")
    parser.add_argument("--rejected", default="quick_users_rechazados.csv")
    args = parser.parse_args()

    run(args.input, args.normalized, args.rejected)


if __name__ == "__main__":
    main()
